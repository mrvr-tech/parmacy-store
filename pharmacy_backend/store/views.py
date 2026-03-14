from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import AllowAny
from django.shortcuts import get_object_or_404, render, redirect
from django.contrib.auth import authenticate
from django.contrib.auth.decorators import login_required
from django.core.exceptions import ValidationError
from rest_framework_simplejwt.tokens import RefreshToken
from .models import StoreItem, LabRequest, User
from .serializers import StoreItemSerializer, LabRequestSerializer, LoginSerializer, UserSerializer
from .decorators import store_keeper_required, lab_user_required
from datetime import datetime, timedelta, date

# ============================================================================
# HELPER FUNCTION: Expiry Alert System
# ============================================================================

def get_expiry_alerts():
    """Get expired and expiring soon items"""
    today = date.today()
    
    expired_items = StoreItem.objects.filter(expiry_date__lt=today).order_by('expiry_date')
    expiring_soon = StoreItem.objects.filter(
        expiry_date__gte=today,
        expiry_date__lte=today + timedelta(days=30)
    ).order_by('expiry_date')
    
    return {
        'expired': list(expired_items),
        'expiring_soon': list(expiring_soon),
        'total_alerts': expired_items.count() + expiring_soon.count()
    }

@api_view(['POST'])
@permission_classes([AllowAny])
def login(request):
    serializer = LoginSerializer(data=request.data)
    
    if serializer.is_valid():
        user = serializer.validated_data['user']
        refresh = RefreshToken.for_user(user)
        
        return Response({
            'refresh': str(refresh),
            'access': str(refresh.access_token),
            'user': UserSerializer(user).data
        })
    
    return Response(serializer.errors, status=400)

@api_view(['GET'])
@permission_classes([AllowAny])
def get_items(request):
    items = StoreItem.objects.all()
    serializer = StoreItemSerializer(items, many=True)
    return Response(serializer.data)

@api_view(['POST'])
@permission_classes([AllowAny])
def add_item(request):
    import logging
    logger = logging.getLogger(__name__)
    
    logger.info(f"Add Item Request Data: {request.data}")
    serializer = StoreItemSerializer(data=request.data)
    
    if serializer.is_valid():
        serializer.save()
        logger.info("Item added successfully")
        return Response({
            "message": "Item added successfully",
            "data": serializer.data
        }, status=201)
    
    logger.error(f"Validation Error: {serializer.errors}")
    return Response({
        "error": "Failed to add item",
        "details": serializer.errors
    }, status=400)


@api_view(['POST'])
@permission_classes([AllowAny])
def create_request(request):
    import logging
    logger = logging.getLogger(__name__)
    
    logger.info(f"Lab Request Data: {request.data}")
    serializer = LabRequestSerializer(data=request.data)
    
    if serializer.is_valid():
        serializer.save()
        logger.info("Lab request created successfully")
        return Response({
            "message": "Lab request created successfully",
            "data": serializer.data
        }, status=201)
    
    logger.error(f"Validation Error: {serializer.errors}")
    return Response({
        "error": "Failed to create lab request",
        "details": serializer.errors
    }, status=400)


@api_view(['POST'])
def approve_request(request, request_id):
    lab_request = get_object_or_404(LabRequest, id=request_id)
    item = lab_request.item
    
    if item.quantity >= lab_request.quantity:
        item.quantity -= lab_request.quantity
        item.save()
        
        lab_request.status = "Approved"
        lab_request.approved_quantity = lab_request.quantity
        lab_request.save()
        
        return Response({"message": "Request Approved"})
    
    return Response({"error": "Not enough stock"}, status=400)


# Dashboard Views - Store Keeper Only
@store_keeper_required
def dashboard(request):
    """Store Keeper dashboard with statistics"""
    total_items = StoreItem.objects.count()
    total_labs = LabRequest.objects.values('lab_name').distinct().count()
    pending_requests = LabRequest.objects.filter(status='Pending').count()
    
    # Low stock items (less than 5)
    low_stock_items = StoreItem.objects.filter(quantity__lt=5).count()
    
    # Expiring soon (within 30 days)
    today = date.today()
    expiring_soon = StoreItem.objects.filter(
        expiry_date__lte=today + timedelta(days=30),
        expiry_date__gte=today
    ).count()
    
    # Get detailed expiry alerts
    expiry_alerts = get_expiry_alerts()
    
    context = {
        'total_items': total_items,
        'total_labs': total_labs,
        'pending_requests': pending_requests,
        'low_stock_items': low_stock_items,
        'expiring_soon': expiring_soon,
        'expired_items': expiry_alerts['expired'],
        'expiring_soon_items': expiry_alerts['expiring_soon'],
        'total_alerts': expiry_alerts['total_alerts'],
    }
    
    return render(request, 'dashboard.html', context)


# Reports Views - Store Keeper Only
@store_keeper_required
def reports(request):
    """Main reports page with all three report sections"""
    items = StoreItem.objects.all().order_by('-date')
    lab_requests = LabRequest.objects.all().order_by('-request_date')
    
    # Calculate totals for each item
    from django.db.models import Sum, Q
    
    stock_report = []
    for item in StoreItem.objects.all():
        used_qty = LabRequest.objects.filter(
            item=item,
            status='Approved'
        ).aggregate(Sum('approved_quantity'))['approved_quantity__sum'] or 0
        
        available = item.quantity - used_qty
        
        stock_report.append({
            'item': item,
            'total_qty': item.quantity,
            'used_qty': used_qty,
            'available_qty': available,
            'expiry_date': item.expiry_date,
        })
    
    context = {
        'items': items,
        'lab_requests': lab_requests,
        'stock_report': stock_report,
    }
    
    return render(request, 'reports.html', context)


@store_keeper_required
def expiry_alerts(request):
    """Dedicated Expiry Alert page showing all expired and expiring items"""
    expiry_alerts = get_expiry_alerts()
    
    context = {
        'expired_items': expiry_alerts['expired'],
        'expiring_soon_items': expiry_alerts['expiring_soon'],
        'total_alerts': expiry_alerts['total_alerts'],
    }
    
    return render(request, 'expiry_alerts.html', context)


@store_keeper_required
def export_purchase_report(request):
    """Export Store Purchase Report as PDF - Store Keeper Only"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from io import BytesIO
    from datetime import datetime
    
    items = StoreItem.objects.all().order_by('-date')
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1a472a'),
        spaceAfter=30,
    )
    
    title = Paragraph("PHARMACY STORE PURCHASE REPORT", title_style)
    elements.append(title)
    
    generated_date = Paragraph(
        f"<b>Generated Date:</b> {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        styles['Normal']
    )
    elements.append(generated_date)
    elements.append(Spacer(1, 0.3*inch))
    
    # Create table data
    table_data = [['Sr.No', 'Item', 'Qty', 'Price (₹)', 'Tax', 'Bill No', 'Date', 'Vendor']]
    
    for idx, item in enumerate(items, 1):
        table_data.append([
            str(idx),
            item.item_name,
            str(item.quantity),
            f"₹{item.price:.2f}",
            f"₹{item.tax:.2f}",
            item.bill_no,
            item.date.strftime('%d/%m/%y'),
            item.vendor_name,
        ])
    
    table = Table(table_data, colWidths=[0.6*inch, 1.5*inch, 0.6*inch, 0.9*inch, 0.6*inch, 0.8*inch, 0.7*inch, 1.5*inch])
    
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a472a')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')]),
    ]))
    
    elements.append(table)
    
    doc.build(elements)
    buffer.seek(0)
    
    from django.http import HttpResponse
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="purchase_report.pdf"'
    return response


@store_keeper_required
def export_lab_usage_report(request):
    """Export Lab Usage Report as PDF"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from io import BytesIO
    from datetime import datetime
    
    lab_requests = LabRequest.objects.all().order_by('-request_date')
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1a472a'),
        spaceAfter=30,
    )
    
    title = Paragraph("PHARMACY LAB USAGE REPORT", title_style)
    elements.append(title)
    
    generated_date = Paragraph(
        f"<b>Generated Date:</b> {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        styles['Normal']
    )
    elements.append(generated_date)
    elements.append(Spacer(1, 0.3*inch))
    
    table_data = [['Lab Name', 'Item', 'Requested Qty', 'Approved Qty', 'Date', 'Status']]
    
    for req in lab_requests:
        table_data.append([
            req.lab_name,
            req.item.item_name,
            str(req.quantity),
            str(req.approved_quantity),
            req.request_date.strftime('%d/%m/%y'),
            req.status,
        ])
    
    table = Table(table_data, colWidths=[1.5*inch, 1.5*inch, 1.2*inch, 1.2*inch, 1*inch, 1*inch])
    
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a472a')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')]),
    ]))
    
    elements.append(table)
    
    doc.build(elements)
    buffer.seek(0)
    
    from django.http import HttpResponse
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="lab_usage_report.pdf"'
    return response


@store_keeper_required
def export_stock_report(request):
    """Export Stock Report as PDF"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from io import BytesIO
    from datetime import datetime
    from django.db.models import Sum
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1a472a'),
        spaceAfter=30,
    )
    
    title = Paragraph("PHARMACY STOCK REPORT", title_style)
    elements.append(title)
    
    generated_date = Paragraph(
        f"<b>Generated Date:</b> {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        styles['Normal']
    )
    elements.append(generated_date)
    elements.append(Spacer(1, 0.3*inch))
    
    table_data = [['Item', 'Total Qty', 'Used Qty', 'Available Qty', 'Expiry Date']]
    
    for item in StoreItem.objects.all():
        used_qty = LabRequest.objects.filter(
            item=item,
            status='Approved'
        ).aggregate(Sum('approved_quantity'))['approved_quantity__sum'] or 0
        
        available = item.quantity - used_qty
        
        table_data.append([
            item.item_name,
            str(item.quantity),
            str(used_qty),
            str(available),
            item.expiry_date.strftime('%d/%m/%y'),
        ])
    
    table = Table(table_data, colWidths=[2*inch, 1.5*inch, 1.5*inch, 1.5*inch, 1.5*inch])
    
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a472a')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')]),
    ]))
    
    elements.append(table)
    
    doc.build(elements)
    buffer.seek(0)
    
    from django.http import HttpResponse
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="stock_report.pdf"'
    return response


@store_keeper_required
def export_purchase_report_excel(request):
    """Export Store Purchase Report as Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.http import HttpResponse
    
    items = StoreItem.objects.all().order_by('-date')
    
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Purchase Report"
    
    # Headers
    headers = ['Sr.No', 'Item Name', 'Quantity', 'Price (₹)', 'Tax (₹)', 'Bill No', 'Date', 'Expiry Date', 'Vendor Name', 'Vendor Address', 'Vendor PAN']
    sheet.append(headers)
    
    # Style header
    header_fill = PatternFill(start_color="1a472a", end_color="1a472a", fill_type="solid")
    header_font = Font(bold=True, color="ffffff")
    
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add data
    for idx, item in enumerate(items, 1):
        sheet.append([
            idx,
            item.item_name,
            item.quantity,
            item.price,
            item.tax,
            item.bill_no,
            item.date.strftime('%d/%m/%Y'),
            item.expiry_date.strftime('%d/%m/%Y'),
            item.vendor_name,
            item.vendor_address,
            item.vendor_pan,
        ])
    
    # Adjust column widths
    sheet.column_dimensions['A'].width = 6
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 12
    sheet.column_dimensions['D'].width = 12
    sheet.column_dimensions['E'].width = 12
    sheet.column_dimensions['F'].width = 12
    sheet.column_dimensions['G'].width = 12
    sheet.column_dimensions['H'].width = 12
    sheet.column_dimensions['I'].width = 20
    sheet.column_dimensions['J'].width = 25
    sheet.column_dimensions['K'].width = 12
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="purchase_report.xlsx"'
    workbook.save(response)
    
    return response


@store_keeper_required
def export_lab_usage_report_excel(request):
    """Export Lab Usage Report as Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse
    
    lab_requests = LabRequest.objects.all().order_by('-request_date')
    
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Lab Usage Report"
    
    headers = ['Lab Name', 'Item Name', 'Requested Quantity', 'Approved Quantity', 'Date', 'Status']
    sheet.append(headers)
    
    header_fill = PatternFill(start_color="1a472a", end_color="1a472a", fill_type="solid")
    header_font = Font(bold=True, color="ffffff")
    
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    for req in lab_requests:
        sheet.append([
            req.lab_name,
            req.item.item_name,
            req.quantity,
            req.approved_quantity,
            req.request_date.strftime('%d/%m/%Y'),
            req.status,
        ])
    
    sheet.column_dimensions['A'].width = 15
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 18
    sheet.column_dimensions['D'].width = 18
    sheet.column_dimensions['E'].width = 12
    sheet.column_dimensions['F'].width = 12
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="lab_usage_report.xlsx"'
    workbook.save(response)
    
    return response


@store_keeper_required
def export_stock_report_excel(request):
    """Export Stock Report as Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.http import HttpResponse
    from django.db.models import Sum
    
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Stock Report"
    
    headers = ['Item Name', 'Total Quantity', 'Used Quantity', 'Available Quantity', 'Expiry Date']
    sheet.append(headers)
    
    header_fill = PatternFill(start_color="1a472a", end_color="1a472a", fill_type="solid")
    header_font = Font(bold=True, color="ffffff")
    
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    for item in StoreItem.objects.all():
        used_qty = LabRequest.objects.filter(
            item=item,
            status='Approved'
        ).aggregate(Sum('approved_quantity'))['approved_quantity__sum'] or 0
        
        available = item.quantity - used_qty
        
        sheet.append([
            item.item_name,
            item.quantity,
            used_qty,
            available,
            item.expiry_date.strftime('%d/%m/%Y'),
        ])
    
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 18
    sheet.column_dimensions['C'].width = 18
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 15
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="stock_report.xlsx"'
    workbook.save(response)
    
    return response


# ============================================================================
# AUTHENTICATION & LOGIN VIEWS
# ============================================================================

def login_view(request):
    """Handle user login for Store Keepers and Lab Users"""
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '').strip()
        
        if not username or not password:
            return render(request, 'login.html', {
                'error': 'Username and password are required.'
            })
        
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            from django.contrib.auth import login
            login(request, user)
            
            # Redirect based on role
            if user.role == 'store':
                return redirect('store_dashboard')
            elif user.role == 'lab':
                return redirect('lab_dashboard')
        else:
            return render(request, 'login.html', {
                'error': 'Invalid username or password.'
            })
    
    return render(request, 'login.html')


@login_required(login_url='login')
def logout_view(request):
    """Handle user logout"""
    from django.contrib.auth import logout
    logout(request)
    return redirect('login')


# ============================================================================
# STORE KEEPER VIEWS (Protected)
# ============================================================================

@store_keeper_required
def store_dashboard(request):
    """Store Keeper Dashboard view"""
    return dashboard(request)


@store_keeper_required
def store_reports(request):
    """Store Keeper Reports view"""
    return reports(request)


@store_keeper_required
def manage_inventory(request):
    """Manage store inventory - view all items"""
    query = request.GET.get('q', '').strip()
    selected_category = request.GET.get('category', '').strip()
    category_lookup = dict(StoreItem.CATEGORY_CHOICES)

    items = StoreItem.objects.all()

    if query:
        items = items.filter(item_name__icontains=query)

    if selected_category in category_lookup:
        items = items.filter(category=selected_category)
    else:
        selected_category = ''

    items = items.order_by('-date')
    
    context = {
        'items': items,
        'search_query': query,
        'selected_category': selected_category,
        'selected_category_label': category_lookup.get(selected_category, ''),
        'category_choices': StoreItem.CATEGORY_CHOICES,
    }
    
    return render(request, 'store/manage_inventory.html', context)


@store_keeper_required
def add_item_form(request):
    """Display add item form"""
    form_data = request.POST.dict() if request.method == 'POST' else {}
    context = {
        'category_choices': StoreItem.CATEGORY_CHOICES,
        'form_data': form_data,
    }

    if request.method == 'POST':
        try:
            from django.db.models import Max

            next_sr_no = (StoreItem.objects.aggregate(max_sr_no=Max('sr_no'))['max_sr_no'] or 0) + 1
            item = StoreItem(
                category=request.POST.get('category', '').strip(),
                sr_no=next_sr_no,
                item_name=request.POST.get('item_name', '').strip(),
                packages=request.POST.get('packages', '').strip() or None,
                quantity=int(request.POST.get('quantity')),
                price=float(request.POST.get('price')),
                tax=float(request.POST.get('tax', 0)),
                bill_no=request.POST.get('bill_no', '').strip(),
                date=request.POST.get('date'),
                expiry_date=request.POST.get('expiry_date'),
                vendor_name=request.POST.get('vendor_name', '').strip(),
                vendor_address=request.POST.get('vendor_address', '').strip(),
                vendor_pan=request.POST.get('vendor_pan', '').strip(),
            )
            item.full_clean()
            item.save()
            return redirect('manage_inventory')
        except (TypeError, ValueError, ValidationError) as e:
            if isinstance(e, ValidationError):
                messages = []
                if hasattr(e, 'message_dict'):
                    for field_messages in e.message_dict.values():
                        messages.extend(field_messages)
                else:
                    messages.extend(e.messages)
                context['error'] = ' '.join(messages)
            else:
                context['error'] = str(e)

            return render(request, 'store/add_item.html', context)
    
    return render(request, 'store/add_item.html', context)


@store_keeper_required
def approve_lab_requests(request):
    """View and approve lab requests"""
    pending_requests = LabRequest.objects.filter(status='Pending').order_by('-request_date')
    approved_requests = LabRequest.objects.filter(status='Approved').order_by('-request_date')
    
    context = {
        'pending_requests': pending_requests,
        'approved_requests': approved_requests,
    }
    
    return render(request, 'store/approve_requests.html', context)


@store_keeper_required
def approve_single_request(request, request_id):
    """Approve a single lab request"""
    lab_request = get_object_or_404(LabRequest, id=request_id)
    item = lab_request.item
    
    if item.quantity >= lab_request.quantity:
        item.quantity -= lab_request.quantity
        item.save()
        
        lab_request.status = 'Approved'
        lab_request.approved_quantity = lab_request.quantity
        lab_request.save()
        
        return redirect('approve_lab_requests')
    else:
        return render(request, 'store/approve_requests.html', {
            'error': f'Not enough stock. Available: {item.quantity}, Requested: {lab_request.quantity}'
        })


@store_keeper_required
def reject_single_request(request, request_id):
    """Reject a lab request"""
    lab_request = get_object_or_404(LabRequest, id=request_id)
    lab_request.status = 'Rejected'
    lab_request.save()
    return redirect('approve_lab_requests')


# ============================================================================
# LAB USER VIEWS (Protected)
# ============================================================================

@lab_user_required
def lab_dashboard(request):
    """Lab User Dashboard - View available items"""
    query = request.GET.get('q', '').strip()
    
    if query:
        items = StoreItem.objects.filter(item_name__icontains=query).order_by('item_name')
    else:
        items = StoreItem.objects.all().order_by('item_name')
    
    # Calculate available quantity for each item
    from django.db.models import Sum
    
    items_with_availability = []
    for item in items:
        used_qty = LabRequest.objects.filter(
            item=item,
            status='Approved'
        ).aggregate(Sum('approved_quantity'))['approved_quantity__sum'] or 0
        
        available = item.quantity - used_qty
        
        items_with_availability.append({
            'item': item,
            'available_qty': available,
            'total_qty': item.quantity,
        })
    
    context = {
        'items': items_with_availability,
        'lab_name': request.user.lab_name,
        'search_query': query,
    }
    
    return render(request, 'lab/dashboard.html', context)


@lab_user_required
def lab_request_item(request):
    """Lab User - Request an item"""
    if request.method == 'POST':
        item_id = request.POST.get('item_id')
        quantity = request.POST.get('quantity')
        
        try:
            item = get_object_or_404(StoreItem, id=item_id)
            
            lab_request = LabRequest(
                lab_name=request.user.lab_name or f"Lab - {request.user.username}",
                item=item,
                quantity=int(quantity),
                status='Pending'
            )
            lab_request.save()
            
            return render(request, 'lab/request_item.html', {
                'items': StoreItem.objects.all(),
                'success': f'Request submitted for {item.item_name}!'
            })
        except Exception as e:
            return render(request, 'lab/request_item.html', {
                'items': StoreItem.objects.all(),
                'error': str(e)
            })
    
    items = StoreItem.objects.all()
    return render(request, 'lab/request_item.html', {'items': items})


@lab_user_required
def lab_request_history(request):
    """Lab User - View their request history"""
    lab_name = request.user.lab_name or f"Lab - {request.user.username}"
    requests = LabRequest.objects.filter(lab_name=lab_name).order_by('-request_date')
    
    context = {
        'requests': requests,
        'lab_name': lab_name,
    }
    
    return render(request, 'lab/request_history.html', context)
